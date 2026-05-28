"""
main.py — Pipeline principal do Playsic.

Converte um arquivo de áudio num chart completo do Clone Hero.

Uso:
    python main.py --audio musica.mp3
    python main.py --audio musica.mp3 --bpm 120.0

Fluxo:
    1. Demucs htdemucs_6s: separa stems → pasta de saída
    2. Extrai album.png dos metadados do áudio original
    3. Gera song.ini definitivo a partir das tags do áudio
    4. Detecta modelos disponíveis (drums + rhythm); roda inferência em paralelo
    5. Consolida xlsx parciais → notes.xlsx unificado
    6. Converte notes.xlsx → notes.mid
    7. Chart em resultados/novas_musicas/charts/<nome>/
    8. Web preview em resultados/novas_musicas/previews/<nome>/
"""
from __future__ import annotations

import argparse
import concurrent.futures
import logging
import shutil
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ── Paths do repositório ──────────────────────────────────────────────────────
REPO_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(REPO_ROOT))
sys.path.insert(0, str(REPO_ROOT / "treinamento"))
sys.path.insert(0, str(REPO_ROOT / "processamento" / "audio"))
sys.path.insert(0, str(REPO_ROOT / "processamento" / "midi_excel"))
sys.path.insert(0, str(REPO_ROOT / "onyx"))

import openpyxl
from mutagen import File as MutagenFile

import song_ini
from separa_audio import separate as separate_audio
import excel_to_midi
import onyx_web_preview

# ── Checkpoints disponíveis ───────────────────────────────────────────────────
_CKPT = REPO_ROOT / "treinamento" / "checkpoint"

AVAILABLE_MODELS: Dict[str, Tuple[Path, Path]] = {
    "drums": (
        _CKPT / "drums" / "drums_crnn_best.pt",
        _CKPT / "drums" / "drums_crnn_meta.pt",
    ),
    "rhythm": (
        _CKPT / "bass" / "bass_crnn_best.pt",
        _CKPT / "bass" / "bass_crnn_meta.pt",
    ),
    "guitar": (
        _CKPT / "guitar" / "guitar_crnn_best.pt",
        _CKPT / "guitar" / "guitar_crnn_meta.pt",
    ),
    "vocals": (
        _CKPT / "vocals" / "vocals_crnn_best.pt",
        _CKPT / "vocals" / "vocals_crnn_meta.pt",
    ),
}

INSTRUMENT_STEMS: Dict[str, str] = {
    "drums":  "drums.ogg",
    "rhythm": "rhythm.ogg",
    "guitar": "guitar.ogg",
    "vocals": "vocals.ogg",
}

CHARTS_DIR   = REPO_ROOT / "resultados" / "novas_musicas" / "charts"
PREVIEWS_DIR = REPO_ROOT / "resultados" / "novas_musicas" / "previews"

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# BPM detection
# ─────────────────────────────────────────────────────────────────────────────
def detect_bpm(audio_path: Path) -> float:
    """Detecta BPM do áudio usando librosa."""
    try:
        import librosa
        y, sr = librosa.load(str(audio_path), sr=None, mono=True)
        tempo, _ = librosa.beat.beat_track(y=y, sr=sr)
        bpm = float(tempo) if not hasattr(tempo, "__len__") else float(tempo[0])
        log.info("BPM detectado: %.2f", bpm)
        return bpm
    except Exception as exc:
        raise RuntimeError(
            f"Falha na detecção automática de BPM: {exc}\n"
            "Use --bpm para fornecer o BPM manualmente."
        ) from exc


# ─────────────────────────────────────────────────────────────────────────────
# Album art extraction
# ─────────────────────────────────────────────────────────────────────────────
def extract_album_art(audio_path: Path, out_png: Path) -> bool:
    """Extrai capa do álbum dos metadados. Retorna True se extraiu com sucesso."""
    try:
        audio = MutagenFile(audio_path)
        if audio is None:
            return False

        # ID3 (MP3) — APIC frame
        if hasattr(audio, "tags") and audio.tags:
            for key in list(audio.tags.keys()):
                if key.startswith("APIC"):
                    data = audio.tags[key].data
                    out_png.write_bytes(data)
                    return True

        # FLAC / OGG com picture block
        if hasattr(audio, "pictures") and audio.pictures:
            out_png.write_bytes(audio.pictures[0].data)
            return True

        # MP4 / M4A — covr atom
        if audio.tags and "covr" in audio.tags:
            covers = audio.tags["covr"]
            if covers:
                out_png.write_bytes(bytes(covers[0]))
                return True

    except Exception as exc:
        log.warning("Não foi possível extrair album art: %s", exc)
    return False


# ─────────────────────────────────────────────────────────────────────────────
# xlsx consolidation
# ─────────────────────────────────────────────────────────────────────────────
def consolidate_xlsx(partial_xlsxes: List[Path], out_xlsx: Path) -> Path:
    """Junta xlsxes parciais (um por instrumento) num único notes.xlsx.

    - A aba 'info' é lida do primeiro arquivo que a contiver e escrita uma só vez.
    - Cada aba de instrumento (drums / rhythm / guitar / vocals) vem do seu xlsx.
    """
    out_wb = openpyxl.Workbook()
    info_written = False

    for xlsx_path in partial_xlsxes:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws_src = wb[sheet_name]
            if sheet_name == "info":
                if not info_written:
                    ws_dst = out_wb.active
                    ws_dst.title = "info"
                    for row in ws_src.iter_rows(values_only=True):
                        ws_dst.append(list(row))
                    info_written = True
            else:
                ws_dst = out_wb.create_sheet(title=sheet_name)
                for row in ws_src.iter_rows(values_only=True):
                    ws_dst.append(list(row))
        wb.close()

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(str(out_xlsx))
    log.info("notes.xlsx consolidado: %s", out_xlsx)
    return out_xlsx


# ─────────────────────────────────────────────────────────────────────────────
# Inferência de um instrumento (executada em thread separada)
# ─────────────────────────────────────────────────────────────────────────────
def _infer_instrument(instrument: str, audio_path: Path, bpm: float,
                      model_path: Path, meta_path: Path,
                      out_xlsx: Path) -> Path:
    import modelo_gera_excel  # importação tardia — evita carregar PyTorch em imports
    modelo_gera_excel.infer(
        audio_path=audio_path,
        bpm=bpm,
        instrument=instrument,
        model_path=model_path,
        meta_path=meta_path,
        out_xlsx=out_xlsx,
    )
    return out_xlsx


# ─────────────────────────────────────────────────────────────────────────────
# Pipeline principal
# ─────────────────────────────────────────────────────────────────────────────
def run_pipeline(audio_path: Path, bpm: Optional[float] = None) -> Path:
    """Executa o pipeline completo. Retorna o path da pasta do chart gerado."""
    audio_path = Path(audio_path).resolve()
    if not audio_path.is_file():
        raise FileNotFoundError(f"Áudio não encontrado: {audio_path}")

    song_name  = audio_path.stem
    chart_dir  = CHARTS_DIR / song_name
    preview_dir = PREVIEWS_DIR / song_name

    print(f"\n{'━' * 60}")
    print(f"  Playsic Pipeline")
    print(f"  Música : {audio_path.name}")
    print(f"  Saída  : {chart_dir}")
    print(f"{'━' * 60}\n")

    # ── 1. Separação de stems ─────────────────────────────────────────────────
    print("[1/5] Separando stems via Demucs...")
    separate_audio(audio_path, CHARTS_DIR)
    if not chart_dir.is_dir():
        raise RuntimeError(f"Demucs não criou a pasta esperada: {chart_dir}")
    print(f"      Stems em: {chart_dir}")

    # ── 2. Album art ──────────────────────────────────────────────────────────
    print("[2/5] Extraindo album art...")
    if extract_album_art(audio_path, chart_dir / "album.png"):
        print("      album.png extraído.")
    else:
        print("      (album art não encontrada nos metadados)")

    # ── 3. song.ini definitivo ────────────────────────────────────────────────
    print("[3/5] Gerando song.ini...")
    song_ini.generate_song_ini(audio_path, chart_dir / "song.ini")
    print("      song.ini gerado.")

    # ── 4. Inferência dos modelos ─────────────────────────────────────────────
    print("[4/5] Inferência dos modelos...")

    if bpm is None:
        print("      BPM não fornecido — detectando automaticamente...")
        bpm_source = chart_dir / "drums.ogg"
        if not bpm_source.exists():
            bpm_source = audio_path
        bpm = detect_bpm(bpm_source)
        print(f"      BPM detectado: {bpm:.2f}")
    else:
        print(f"      BPM: {bpm:.2f}")

    temp_dir = chart_dir / "_temp"
    temp_dir.mkdir(exist_ok=True)

    active: List[Tuple[str, Path, Path]] = []
    for instrument, (model_path, meta_path) in AVAILABLE_MODELS.items():
        if model_path.exists() and meta_path.exists():
            active.append((instrument, model_path, meta_path))
            print(f"      Modelo disponível : {instrument}")
        else:
            print(f"      Modelo ausente (pulando): {instrument}")

    partial_xlsxes: List[Path] = []

    if active:
        futures: Dict = {}
        with concurrent.futures.ThreadPoolExecutor() as executor:
            for instrument, model_path, meta_path in active:
                stem_name  = INSTRUMENT_STEMS.get(instrument, f"{instrument}.ogg")
                audio_stem = chart_dir / stem_name
                if not audio_stem.exists():
                    print(f"      [AVISO] {stem_name} não encontrado; pulando {instrument}")
                    continue
                out_xlsx = temp_dir / f"{instrument}.xlsx"
                fut = executor.submit(
                    _infer_instrument,
                    instrument, audio_stem, bpm,
                    model_path, meta_path, out_xlsx,
                )
                futures[fut] = instrument

        for fut, instrument in futures.items():
            try:
                xlsx = fut.result()
                partial_xlsxes.append(xlsx)
                print(f"      ✔ {instrument}: {xlsx.name}")
            except Exception as exc:
                print(f"      ✘ {instrument} falhou: {exc}")
    else:
        print("      [AVISO] Nenhum modelo disponível. notes.mid não será gerado.")

    # ── 5. Consolidar xlsx → notes.mid ────────────────────────────────────────
    if partial_xlsxes:
        print("[5/5] Gerando notes.mid...")
        notes_xlsx = temp_dir / "notes.xlsx"
        consolidate_xlsx(partial_xlsxes, notes_xlsx)
        excel_to_midi.convert(notes_xlsx, chart_dir / "notes.mid")
        print("      notes.mid gerado.")
    else:
        print("[5/5] Nenhum xlsx parcial — notes.mid não gerado.")

    # Limpa pasta temporária
    shutil.rmtree(temp_dir, ignore_errors=True)

    print(f"\n{'━' * 60}")
    print(f"  Chart pronto: {chart_dir}")
    print(f"  Arquivos: {sorted(p.name for p in chart_dir.iterdir())}")

    # ── Preview ───────────────────────────────────────────────────────────────
    print("\n[+] Gerando web preview...")
    preview_dir.mkdir(parents=True, exist_ok=True)
    try:
        html = onyx_web_preview.build_preview(chart_dir, preview_dir)
        print(f"  Preview: {html}")
    except Exception as exc:
        print(f"  ✘ Preview não gerado: {exc}")
        print(f"     (execute onyx_web_preview.py manualmente sobre {chart_dir})")

    print(f"{'━' * 60}\n")
    return chart_dir


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Playsic — converte um arquivo de áudio num chart do Clone Hero.",
    )
    p.add_argument(
        "--audio", required=True, type=Path,
        help="Caminho do arquivo de áudio (mp3, flac, ogg, ...)",
    )
    p.add_argument(
        "--bpm", type=float, default=None,
        help="BPM da música (detectado automaticamente via librosa se omitido)",
    )
    return p


def main(argv: Optional[list] = None) -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%H:%M:%S",
    )
    args = _build_arg_parser().parse_args(argv)
    run_pipeline(args.audio, bpm=args.bpm)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

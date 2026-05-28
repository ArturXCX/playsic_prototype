"""
midi_to_excel — converte um arquivo .mid para .xlsx detalhado (com formatação).

Este é o formato "detalhado" usado para análise manual / debug de arquivos
MIDI, NÃO o formato resumido consumido pelo pipeline de inferência. O formato
resumido é gerado internamente pelo `dados/organizar_dataset.py`.

Estrutura do .xlsx detalhado:
    - Aba 'INFO' (maiúsculo): metadados do MIDI + resumo das tracks
    - Uma aba por track no formato 'PART DRUMS (0)', 'PART GUITAR (0)', etc.
      contendo todas as notas + seções de Text Events e Lyrics

Uso (CLI):
    python midi_to_excel.py --midi notes.mid --out notes_detail.xlsx

Uso (API):
    from processamento.midi_excel.midi_to_excel import convert
    convert("notes.mid", "notes_detail.xlsx")
"""
from __future__ import annotations

import argparse
import os
from pathlib import Path
from typing import Any, Dict, List, Optional

import mido
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ─── cores do tema visual ───────────────────────────────────────────────────
HEADER_BG   = "1F4E79"
HEADER_FG   = "FFFFFF"
SUB_BG      = "2E75B6"
SUB_FG      = "FFFFFF"
ALT_ROW_BG  = "D6E4F0"
INFO_KEY_BG = "2E75B6"
INFO_KEY_FG = "FFFFFF"
INFO_VAL_BG = "EBF3FB"

NOTE_NAMES = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]


# ─────────────────────────────────────────────────────────────────────────────
# helpers
# ─────────────────────────────────────────────────────────────────────────────
def _note_name(n: int) -> str:
    return f"{NOTE_NAMES[n % 12]}{(n // 12) - 1}"


def _ticks_to_seconds(ticks: int, tempo: int, tpb: int) -> float:
    return (ticks / tpb) * (tempo / 1_000_000)


def _header_cell(ws, row: int, col: int, value: Any, bg: str = HEADER_BG,
                 fg: str = HEADER_FG, bold: bool = True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c


def _info_key(ws, row: int, col: int, value: Any):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, color=INFO_KEY_FG, size=10)
    c.fill = PatternFill("solid", start_color=INFO_KEY_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c


def _info_val(ws, row: int, col: int, value: Any):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10)
    c.fill = PatternFill("solid", start_color=INFO_VAL_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c


def _data_row(ws, row_num: int, values: List[Any], alt: bool = False) -> None:
    bg = ALT_ROW_BG if alt else "FFFFFF"
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.font = Font(name="Arial", size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(vertical="center")


# ─────────────────────────────────────────────────────────────────────────────
# parsing
# ─────────────────────────────────────────────────────────────────────────────
def parse_midi(midi_path: Path | str) -> Dict[str, Any]:
    mid = mido.MidiFile(str(midi_path))
    tpb = mid.ticks_per_beat
    tempo = 500_000
    time_sig = (4, 4)

    for msg in mid.tracks[0]:
        if msg.type == "set_tempo":
            tempo = msg.tempo
        elif msg.type == "time_signature":
            time_sig = (msg.numerator, msg.denominator)

    tracks_data: List[Dict[str, Any]] = []
    for track in mid.tracks[1:]:
        notes, texts, lyrics = [], [], []
        abs_tick = 0
        open_notes: Dict[int, tuple] = {}
        for msg in track:
            abs_tick += msg.time
            t_sec = _ticks_to_seconds(abs_tick, tempo, tpb)
            if msg.type == "note_on":
                if msg.velocity > 0:
                    open_notes[msg.note] = (abs_tick, t_sec, msg.velocity, msg.channel)
                else:
                    if msg.note in open_notes:
                        s_tick, s_sec, vel, ch = open_notes.pop(msg.note)
                        dur_ticks = abs_tick - s_tick
                        dur_sec = _ticks_to_seconds(dur_ticks, tempo, tpb)
                        notes.append({
                            "note": msg.note,
                            "note_name": _note_name(msg.note),
                            "channel": ch,
                            "velocity": vel,
                            "start_tick": s_tick,
                            "start_sec": round(s_sec, 4),
                            "end_tick": abs_tick,
                            "end_sec": round(s_sec + dur_sec, 4),
                            "duration_ticks": dur_ticks,
                            "duration_sec": round(dur_sec, 4),
                        })
            elif msg.type == "text":
                texts.append({"tick": abs_tick, "time_sec": round(t_sec, 4), "text": msg.text})
            elif msg.type == "lyrics":
                lyrics.append({"tick": abs_tick, "time_sec": round(t_sec, 4), "text": msg.text})
        tracks_data.append({"name": track.name, "notes": notes, "texts": texts, "lyrics": lyrics})

    return {
        "filename":       os.path.basename(str(midi_path)),
        "type":           mid.type,
        "ticks_per_beat": tpb,
        "tempo_us":       tempo,
        "bpm":            round(60_000_000 / tempo, 2),
        "time_signature": f"{time_sig[0]}/{time_sig[1]}",
        "num_tracks":     len(mid.tracks) - 1,
        "tracks":         tracks_data,
    }


# ─────────────────────────────────────────────────────────────────────────────
# escrita do .xlsx
# ─────────────────────────────────────────────────────────────────────────────
def _build_info_sheet(wb: openpyxl.Workbook, data: Dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "INFO"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"MIDI File Info — {data['filename']}"
    c.font = Font(name="Arial", bold=True, size=14, color=HEADER_FG)
    c.fill = PatternFill("solid", start_color=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = [
        ("File Name",        data["filename"]),
        ("MIDI Type",        f"Type {data['type']}"),
        ("Ticks per Beat",   data["ticks_per_beat"]),
        ("Tempo (µs/beat)",  data["tempo_us"]),
        ("BPM",              data["bpm"]),
        ("Time Signature",   data["time_signature"]),
        ("Number of Tracks", data["num_tracks"]),
    ]
    for i, (k, v) in enumerate(meta, 3):
        _info_key(ws, i, 1, k)
        _info_val(ws, i, 2, v)
        ws.row_dimensions[i].height = 18

    r = 3 + len(meta) + 2
    ws.merge_cells(f"A{r}:E{r}")
    c = ws.cell(row=r, column=1, value="Tracks Summary")
    c.font = Font(name="Arial", bold=True, size=11, color=HEADER_FG)
    c.fill = PatternFill("solid", start_color=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22

    r += 1
    for ci, col in enumerate(["#", "Track Name", "Notes", "Text Events", "Lyrics"], 1):
        _header_cell(ws, r, ci, col, bg=SUB_BG)
    ws.row_dimensions[r].height = 18
    r += 1
    for i, t in enumerate(data["tracks"]):
        _data_row(ws, r,
                  [i + 1, f"{t['name']} (0)", len(t["notes"]), len(t["texts"]), len(t["lyrics"])],
                  alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 16
        r += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12


def _build_track_sheet(wb: openpyxl.Workbook, track: Dict[str, Any], idx: int) -> None:
    name = f"{track['name']} (0)"
    ws = wb.create_sheet(title=name[:31])
    ws.sheet_view.showGridLines = False
    total_cols = 12
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value = f"Track: {name}"
    c.font = Font(name="Arial", bold=True, size=13, color=HEADER_FG)
    c.fill = PatternFill("solid", start_color=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    notes  = track["notes"]
    texts  = track["texts"]
    lyrics = track["lyrics"]

    r = 3
    ws.merge_cells(f"A{r}:{get_column_letter(total_cols)}{r}")
    c = ws.cell(row=r, column=1, value=f"Notes  ({len(notes)} total)")
    c.font = Font(name="Arial", bold=True, size=11, color=SUB_FG)
    c.fill = PatternFill("solid", start_color=SUB_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

    note_cols = ["#", "Note #", "Note Name", "Channel", "Velocity",
                 "Start Tick", "Start (s)", "End Tick", "End (s)",
                 "Duration (ticks)", "Duration (s)", "Lyrics / Text"]
    for ci, col in enumerate(note_cols, 1):
        _header_cell(ws, r, ci, col)
    ws.row_dimensions[r].height = 18
    r += 1

    lyric_map: Dict[int, List[str]] = {}
    for lyr in lyrics:
        lyric_map.setdefault(lyr["tick"], []).append(lyr["text"])
    for txt in texts:
        lyric_map.setdefault(txt["tick"], []).append(txt["text"])

    for i, n in enumerate(notes):
        lyric = " | ".join(lyric_map.get(n["start_tick"], []))
        row_vals = [
            i + 1, n["note"], n["note_name"], n["channel"], n["velocity"],
            n["start_tick"], n["start_sec"], n["end_tick"], n["end_sec"],
            n["duration_ticks"], n["duration_sec"], lyric,
        ]
        _data_row(ws, r, row_vals, alt=(i % 2 == 1))
        ws.row_dimensions[r].height = 15
        r += 1

    if texts:
        r += 1
        ws.merge_cells(f"A{r}:C{r}")
        c = ws.cell(row=r, column=1, value=f"Text Events  ({len(texts)} total)")
        c.font = Font(name="Arial", bold=True, size=11, color=SUB_FG)
        c.fill = PatternFill("solid", start_color=SUB_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1
        for ci, col in enumerate(["Tick", "Time (s)", "Text"], 1):
            _header_cell(ws, r, ci, col)
        ws.row_dimensions[r].height = 18
        r += 1
        for i, t in enumerate(texts):
            _data_row(ws, r, [t["tick"], t["time_sec"], t["text"]], alt=(i % 2 == 1))
            ws.row_dimensions[r].height = 15
            r += 1

    if lyrics:
        r += 1
        ws.merge_cells(f"A{r}:C{r}")
        c = ws.cell(row=r, column=1, value=f"Lyrics  ({len(lyrics)} total)")
        c.font = Font(name="Arial", bold=True, size=11, color=SUB_FG)
        c.fill = PatternFill("solid", start_color=SUB_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1
        for ci, col in enumerate(["Tick", "Time (s)", "Lyric"], 1):
            _header_cell(ws, r, ci, col)
        ws.row_dimensions[r].height = 18
        r += 1
        for i, lyr in enumerate(lyrics):
            _data_row(ws, r, [lyr["tick"], lyr["time_sec"], lyr["text"]], alt=(i % 2 == 1))
            ws.row_dimensions[r].height = 15
            r += 1

    col_widths = [6, 8, 12, 9, 10, 14, 11, 14, 11, 18, 14, 30]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# API pública
# ─────────────────────────────────────────────────────────────────────────────
def convert(midi_path: Path | str, out_xlsx: Path | str) -> Path:
    """Converte .mid em .xlsx detalhado (formatado, com abas INFO + por track)."""
    midi_path = Path(midi_path)
    out_xlsx  = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    data = parse_midi(midi_path)
    wb = openpyxl.Workbook()
    _build_info_sheet(wb, data)
    for i, track in enumerate(data["tracks"]):
        _build_track_sheet(wb, track, i)
    wb.save(out_xlsx)
    return out_xlsx


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Converte um arquivo MIDI em planilha Excel detalhada (formato análise).",
    )
    p.add_argument("--midi", required=True, type=Path, help="Arquivo .mid de entrada")
    p.add_argument("--out",  required=True, type=Path, help="Arquivo .xlsx de saída")
    return p


def main(argv: Optional[list] = None) -> int:
    args = _build_arg_parser().parse_args(argv)
    out = convert(args.midi, args.out)
    print(f"xlsx gerado em: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

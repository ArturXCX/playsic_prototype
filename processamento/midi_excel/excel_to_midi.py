"""
excel_to_midi — converte .xlsx (formato resumido) → .mid.

Formato resumido esperado:
    - Aba 'info' (minúsculo): 1 linha de cabeçalho + 1 linha de dados.
      Colunas: File Name, MIDI Type, Ticks per Beat, Tempo (µs/beat),
      BPM, Time Signature.
    - Uma ou mais abas de instrumento: 'guitar', 'drums', 'rhythm', 'vocals'.
      Cabeçalho com 11 colunas:
        # | Note # | Note Name | Channel | Velocity |
        Start Tick | Start (s) | End Tick | End (s) |
        Duration (ticks) | Duration (s)

Mapeamento aba → nome de track MIDI:
    guitar → PART GUITAR
    drums  → PART DRUMS
    rhythm → PART BASS
    vocals → PART VOCALS

Abas desconhecidas são ignoradas (com aviso).

Uso (CLI):
    python excel_to_midi.py --xlsx notes.xlsx --out notes.mid

Uso (API):
    from processamento.midi_excel.excel_to_midi import convert
    convert("notes.xlsx", "notes.mid")
"""
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional

import mido
import openpyxl
from mido import MetaMessage, Message, MidiFile, MidiTrack


SHEET_TO_TRACK = {
    "guitar": "PART GUITAR",
    "drums":  "PART DRUMS",
    "rhythm": "PART BASS",
    "vocals": "PART VOCALS",
}

# ─────────────────────────────────────────────────────────────────────────────
# Remapeamento de notas de drums para Clone Hero (4 dificuldades)
# ─────────────────────────────────────────────────────────────────────────────
# O modelo usa as notas de anotação (24=Kick,26=Snare,27=Yellow,30=Blue,31=Green).
# Convertemos para a convenção OFICIAL de drums do Clone Hero/YARG (4-lane), em que
# CADA dificuldade tem: Kick, Red(snare), Yellow, Blue, Green em pitches consecutivos:
#   Easy 60-64 | Medium 72-76 | Hard 84-88 | Expert 96-100
#   (Kick=60/72/84/96, Snare=61/73/85/97, Yellow=62/74/86/98, Blue=63/75/87/99, Green=64/76/88/100)
# Antes o kick ficava na nota 24 (fora do range jogável) e os pads estavam deslocados.
_DRUMS_PAD_TO_DIFFICULTIES: Dict[int, List[int]] = {
    24: [60, 72, 84, 96],    # Kick
    26: [61, 73, 85, 97],    # Snare (Red)
    27: [62, 74, 86, 98],    # Yellow (hi-hat)
    30: [63, 75, 87, 99],    # Blue (tom)
    31: [64, 76, 88, 100],   # Green (crash)
}

# Guitar / Bass (rhythm): o modelo é treinado nas notas Expert (96-100).
# Clone Hero exige notas em cada faixa de dificuldade para que o chart seja
# jogável. Expandimos cada nota Expert para os 4 níveis.
#   Easy 60-64 | Medium 72-76 | Hard 84-88 | Expert 96-100
_RHYTHM_FRET_TO_DIFFICULTIES: Dict[int, List[int]] = {
    96:  [60, 72, 84, 96],   # Green:  Easy → Expert
    97:  [61, 73, 85, 97],   # Red:    Easy → Expert
    98:  [62, 74, 86, 98],   # Yellow: Easy → Expert
    99:  [63, 75, 87, 99],   # Blue:   Easy → Expert
    100: [64, 76, 88, 100],  # Orange: Easy → Expert
}

# Guitar lead: mesma estrutura do rhythm (5 frets, mesmas notas Expert).
_GUITAR_FRET_TO_DIFFICULTIES: Dict[int, List[int]] = dict(_RHYTHM_FRET_TO_DIFFICULTIES)


# ─────────────────────────────────────────────────────────────────────────────
# Redução por dificuldade (Easy < Medium < Hard < Expert)
# ─────────────────────────────────────────────────────────────────────────────
# O esquema antigo COPIAVA o Expert idêntico para os 4 níveis — Easy ficava tão
# denso quanto Expert (injogável). Um chart de verdade (e o que o Onyx faz na
# conversão) gera cada nível inferior como um SUBCONJUNTO afinado do Expert:
# menos notas (espaçamento rítmico maior) + acordes simplificados.
#
# índice 0=Easy, 1=Medium, 2=Hard, 3=Expert  →  (espaçamento mínimo em beats,
# nº máx. de notas simultâneas).
_DIFFICULTY_PARAMS = [
    (0, 1.5,   1),   # Easy:   >= 1.5 beats entre notas, sem acordes
    (1, 0.75,  1),   # Medium: >= 3/4 de beat, sem acordes
    (2, 0.375, 2),   # Hard:   >= 3/8 de beat (remove 16ths/tercinas), até 2 notas
    (3, 0.0,   99),  # Expert: tudo (idêntico à entrada)
]

# Prioridade ao cortar acordes nos níveis baixos (menor = mantido primeiro).
# Guitar/bass: mantém o fret mais grave (nota Expert menor = Green/Red).
def _fret_chord_priority(expert_note: int) -> int:
    return expert_note

# Drums: mantém o backbone (kick → caixa → chimbal → tom → prato) ao simplificar.
_DRUM_PRIORITY = {24: 0, 26: 1, 27: 2, 30: 3, 31: 4}
def _drum_chord_priority(expert_note: int) -> int:
    return _DRUM_PRIORITY.get(expert_note, 9)


def _reduce_to_difficulties(notes: List[Dict[str, Any]],
                            ticks_per_beat: int,
                            expert_map: Dict[int, List[int]],
                            chord_priority) -> List[Dict[str, Any]]:
    """Gera as 4 dificuldades a partir das notas Expert, com densidade decrescente.

    Args:
        notes:          notas no range Expert (ex.: 96-100 p/ frets, 26-31 p/ pads)
        ticks_per_beat: resolução do MIDI
        expert_map:     {nota_expert: [easy, medium, hard, expert]} (alvo por nível)
        chord_priority: callable(nota_expert) -> int (ordem de corte de acordes)

    Notas fora de `expert_map` (ex.: kick 24) passam inalteradas, uma única vez.
    Expert é idêntico à entrada; níveis inferiores são subconjuntos afinados.
    """
    import collections

    out: List[Dict[str, Any]] = []
    mapped = [n for n in notes if n["note"] in expert_map]
    # passthrough: notas fora do mapa de dificuldade (kick, marcadores, etc.)
    for n in notes:
        if n["note"] not in expert_map:
            out.append(n)

    by_tick: Dict[int, List[Dict[str, Any]]] = collections.defaultdict(list)
    for n in mapped:
        by_tick[n["start_tick"]].append(n)
    ticks = sorted(by_tick)

    for di, gap_beats, max_chord in _DIFFICULTY_PARAMS:
        min_gap = round(gap_beats * ticks_per_beat)
        last = None
        for tick in ticks:
            if di < 3 and last is not None and (tick - last) < min_gap:
                continue                      # afina: pula notas muito próximas
            last = tick
            chord = sorted(by_tick[tick], key=lambda n: chord_priority(n["note"]))
            for n in chord[:max_chord]:       # simplifica acordes nos níveis baixos
                out.append({**n, "note": expert_map[n["note"]][di]})
    return out


# ─────────────────────────────────────────────────────────────────────────────
# leitura do .xlsx resumido
# ─────────────────────────────────────────────────────────────────────────────
def read_info(wb: openpyxl.Workbook) -> Dict[str, int]:
    """Lê a aba 'info' e retorna metadados globais."""
    ws = wb["info"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    values  = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    row = dict(zip(headers, values))

    ts = str(row.get("Time Signature", "4/4"))
    num, den = ts.split("/")

    return {
        "midi_type":      int(str(row.get("MIDI Type", "Type 1")).replace("Type ", "")),
        "ticks_per_beat": int(row.get("Ticks per Beat", 480)),
        "tempo":          int(row.get("Tempo (µs/beat)", 500_000)),
        "time_sig_num":   int(num),
        "time_sig_den":   int(den),
    }


def read_track_sheet(ws) -> List[Dict[str, int]]:
    """Lê uma aba de instrumento e retorna a lista de notas."""
    notes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        note_num = row[1]
        if note_num is None:
            continue
        channel  = row[3]
        velocity = row[4]
        start_tk = row[5]
        end_tk   = row[7]
        dur_tk   = row[9]
        notes.append({
            "note":       int(note_num),
            "channel":    int(channel)  if channel  is not None else 0,
            "velocity":   int(velocity) if velocity is not None else 64,
            "start_tick": int(start_tk),
            "end_tick":   int(end_tk) if end_tk is not None
                          else int(start_tk) + int(dur_tk),
        })
    return notes


# ─────────────────────────────────────────────────────────────────────────────
# montagem do .mid
# ─────────────────────────────────────────────────────────────────────────────
def build_midi(wb: openpyxl.Workbook, out_path: Path) -> Dict[str, Any]:
    """Monta o .mid a partir de um workbook aberto. Retorna estatísticas."""
    info  = read_info(wb)
    tpb   = info["ticks_per_beat"]
    tempo = info["tempo"]

    mid = MidiFile(type=1, ticks_per_beat=tpb)
    t0 = MidiTrack()
    mid.tracks.append(t0)
    t0.append(MetaMessage("track_name", name="notes", time=0))
    t0.append(MetaMessage("set_tempo",  tempo=tempo, time=0))
    t0.append(MetaMessage(
        "time_signature",
        numerator=info["time_sig_num"],
        denominator=info["time_sig_den"],
        clocks_per_click=24,
        notated_32nd_notes_per_beat=8,
        time=0,
    ))
    t0.append(MetaMessage("end_of_track", time=0))

    TYPE_ORDER = {"note_off": 0, "note_on": 2}
    track_counts: Dict[str, int] = {}
    ignored: List[str] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "info":
            continue
        track_name = SHEET_TO_TRACK.get(sheet_name)
        if track_name is None:
            ignored.append(sheet_name)
            continue

        notes = read_track_sheet(wb[sheet_name])

        # Expansão por dificuldade com REDUÇÃO real (Easy < Medium < Hard < Expert).
        # As notas vêm no range Expert; cada nível inferior é um subconjunto
        # afinado (menos notas + acordes simplificados), como um chart de verdade —
        # não uma cópia do Expert em outra faixa de notas. Ver _reduce_to_difficulties.
        if sheet_name == "drums":
            notes = _reduce_to_difficulties(
                notes, tpb, _DRUMS_PAD_TO_DIFFICULTIES, _drum_chord_priority)

        elif sheet_name == "rhythm":
            notes = _reduce_to_difficulties(
                notes, tpb, _RHYTHM_FRET_TO_DIFFICULTIES, _fret_chord_priority)

        elif sheet_name == "guitar":
            notes = _reduce_to_difficulties(
                notes, tpb, _GUITAR_FRET_TO_DIFFICULTIES, _fret_chord_priority)

        # Vocals (v1): o modelo gera onsets como nota MIDI 60. Não expandimos
        # — vocais no Clone Hero usam PART VOCALS com pitches reais, não por
        # dificuldade. Mantemos as notas como vieram.

        events = []
        for n in notes:
            events.append((n["start_tick"], "note_on",  n))
            events.append((n["end_tick"],   "note_off", n))
        events.sort(key=lambda e: (e[0], TYPE_ORDER.get(e[1], 1)))

        tr = MidiTrack()
        mid.tracks.append(tr)
        tr.append(MetaMessage("track_name", name=track_name, time=0))
        prev = 0
        for abs_tick, etype, n in events:
            delta = abs_tick - prev
            prev  = abs_tick
            vel = n["velocity"] if etype == "note_on" else 0
            tr.append(Message(
                "note_on",
                channel=n["channel"],
                note=n["note"],
                velocity=vel,
                time=delta,
            ))
        tr.append(MetaMessage("end_of_track", time=0))
        track_counts[track_name] = len(notes)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    mid.save(str(out_path))
    return {
        "midi_path":     out_path,
        "ticks_per_beat": tpb,
        "tempo_us":       tempo,
        "bpm":            round(60_000_000 / tempo, 2),
        "track_counts":   track_counts,
        "ignored_sheets": ignored,
    }


# ─────────────────────────────────────────────────────────────────────────────
# API pública
# ─────────────────────────────────────────────────────────────────────────────
def convert(xlsx_path: Path | str, out_mid: Path | str) -> Path:
    """Converte um .xlsx resumido em .mid e retorna o Path do arquivo gerado."""
    xlsx_path = Path(xlsx_path)
    out_mid   = Path(out_mid)
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    result = build_midi(wb, out_mid)
    wb.close()
    return result["midi_path"]


# ─────────────────────────────────────────────────────────────────────────────
# verificação opcional
# ─────────────────────────────────────────────────────────────────────────────
def verify(original_mid: Path | str, rebuilt_mid: Path | str) -> bool:
    """Compara contagem de notas por track entre o MIDI original e o reconstruído.

    Retorna True se todas as tracks do reconstruído batem com o original.
    Imprime uma tabela para inspeção visual.
    """
    orig = mido.MidiFile(str(original_mid))
    new  = mido.MidiFile(str(rebuilt_mid))

    def count(midfile):
        return [(t.name, sum(1 for m in t if m.type == "note_on" and m.velocity > 0))
                for t in midfile.tracks[1:]]

    orig_counts = count(orig)
    new_counts  = dict(count(new))

    print(f"\n{'Track':<22} {'Original':>10} {'Reconstruído':>14} {'OK?':>5}")
    print("-" * 55)
    all_ok = True
    for name, oc in orig_counts:
        nc = new_counts.get(name)
        if nc is None:
            print(f"{name:<22} {oc:>10} {'(ausente)':>14}")
            continue
        ok = "✓" if oc == nc else "✗"
        if oc != nc:
            all_ok = False
        print(f"{name:<22} {oc:>10} {nc:>14} {ok:>5}")
    print("-" * 55)
    print("✓ Tudo bate!" if all_ok else "✗ Há diferenças — revise.")
    return all_ok


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Converte uma planilha Excel resumida em arquivo MIDI.",
    )
    p.add_argument("--xlsx",    required=True, type=Path, help="Planilha de entrada (formato resumido)")
    p.add_argument("--out",     required=True, type=Path, help="Arquivo .mid de saída")
    p.add_argument("--verify",  type=Path, default=None,
                   help="Caminho de um .mid original. Se informado, compara notas por track.")
    return p


def main(argv: Optional[list] = None) -> int:
    args = _build_arg_parser().parse_args(argv)
    out  = convert(args.xlsx, args.out)
    print(f"MIDI salvo em: {out}")
    if args.verify is not None:
        verify(args.verify, out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

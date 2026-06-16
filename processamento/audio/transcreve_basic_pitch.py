"""
transcreve_basic_pitch — transcrição áudio→notas via basic-pitch (backend ONNX),
mapeada para o xlsx "resumido" do Clone Hero.

É uma alternativa ao CRNN (`treinamento/modelo_gera_excel.infer`) para os
instrumentos AFINADOS — guitar, bass/rhythm e vocals — com a MESMA assinatura de
saída, então entra como drop-in no `main.py`. Drums NÃO é suportado (basic-pitch
não transcreve bateria): use o CRNN para drums.

Por que basic-pitch:
    O CRNN tinha que adivinhar o chart Expert direto do áudio, num grid de
    semicolcheia — tarefa dificílima com dataset pequeno. basic-pitch resolve a
    parte de PERCEPÇÃO (áudio → notas com onset+pitch+duração) com um modelo
    pré-treinado robusto; aqui só mapeamos as notas para as lanes do jogo.

Mapeamento pitch → fret (guitar/bass): CONTORNO MELÓDICO.
    O fret acompanha o movimento da melodia: subiu de tom → fret mais alto,
    desceu → mais baixo (relativo à nota anterior, ~1 fret por tom). Gera charts
    ergonômicos que "sobem e descem" junto com o solo, em vez de grudar num fret.
    Monofônico na v1 (1 nota por step; em acordes mantém a de maior amplitude).

Vocals: PART VOCALS recebe os PITCHES REAIS cantados (grande melhoria sobre o
    monotone v1 do CRNN), com durações reais, fixados na faixa vocal [36, 84].

Backend ONNX (onnxruntime) — não usa TensorFlow, não mexe no ambiente de treino.

Uso (API):
    from processamento.audio.transcreve_basic_pitch import transcribe
    transcribe("guitar.ogg", bpm=120.0, instrument="guitar", out_xlsx="guitar.xlsx")

Uso (CLI):
    python transcreve_basic_pitch.py --audio guitar.ogg --bpm 120 \
        --instrument guitar --out guitar.xlsx
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np

# ── Imports irmãos (audio_features + notes_xlsx vivem em treinamento/) ─────────
_HERE = Path(__file__).resolve()
_REPO = _HERE.parents[2]
sys.path.insert(0, str(_HERE.parent))               # processamento/audio (onset_model)
sys.path.insert(0, str(_REPO / "treinamento"))
from audio_features import (  # noqa: E402
    SUBDIV_PER_BEAT, TICKS_PER_BEAT, audio_duration_seconds, step_duration_seconds,
)
from notes_xlsx import midi_to_note_name, predictions_to_xlsx  # noqa: E402


log = logging.getLogger(__name__)

SUPPORTED_INSTRUMENTS = {"guitar", "rhythm", "vocals"}

# ── Mapeamento de frets ───────────────────────────────────────────────────────
N_FRETS        = 5
CONTOUR_WINDOW = 12             # nº de notas recentes que definem o registro local
START_FRET     = 2             # fret inicial (Yellow), até a janela encher
# fret_idx → nota MIDI Expert (Green..Orange). excel_to_midi expande/reduz daqui.
FRET_IDX_TO_MIDI: Dict[int, int] = {0: 96, 1: 97, 2: 98, 3: 99, 4: 100}

# ── Densidade-alvo (afina onsets para ~densidade do chart humano) ─────────────
# Mantém os onsets metricamente mais fortes até o cap de onsets/seg; músicas
# abaixo do cap ficam inalteradas. Baixo (bass) é menos denso que guitarra.
MAX_ONSETS_PER_SEC = {"guitar": 3.0, "rhythm": 1.5}

# ── Acordes (polifonia do basic-pitch) ───────────────────────────────────────
# Nº máx. de frets simultâneos por instrumento. Bass é ~monofônico em charts
# reais (chord_rate ~1%), então 1; guitar usa até 3.
MAX_CHORD      = {"guitar": 3, "rhythm": 1}
CHORD_AMP_FRAC = 0.4    # tom do acorde só conta se amp >= fração da amp do líder

# ── Sustains ──────────────────────────────────────────────────────────────────
# Charts humanos são ~92% taps curtos + ~8% notas longas (sustains). Emitimos
# sustain só nas notas mais longas de cada música (limiar = percentil por-música),
# casando a estatística humana, com um piso absoluto. As demais viram tap curto.
SUSTAIN_RATE      = 0.10  # fração de notas que viram sustain (alvo ~humano)
SUSTAIN_MIN_STEPS = 4     # piso: nunca sustenta nota < 1 beat (4 semicolcheias)
TAP_TICKS         = 60    # duração de nota curta (tap), convenção Clone Hero

# ── Faixa vocal considerada (fora disso é descartado/fixado) ──────────────────
VOCAL_MIDI_MIN = 36
VOCAL_MIDI_MAX = 84


# ─────────────────────────────────────────────────────────────────────────────
# basic-pitch (ONNX)
# ─────────────────────────────────────────────────────────────────────────────
def _onnx_model_path() -> Path:
    """Caminho do modelo ONNX que vem junto do pacote basic-pitch."""
    from basic_pitch import ICASSP_2022_MODEL_PATH
    return Path(str(ICASSP_2022_MODEL_PATH)).parent / "nmp.onnx"


# Sensibilidade do basic-pitch. Abaixe ONSET/FRAME para detectar MAIS notas
# (stems quietos/distorcidos); MIN_NOTE_LEN_MS menor mantém notas mais curtas.
ONSET_THRESHOLD = 0.5     # default basic-pitch; ↓ = mais onsets
FRAME_THRESHOLD = 0.3     # default basic-pitch; ↓ = sustains mais longos/sensível
MIN_NOTE_LEN_MS = 100.0   # default 127.7; ↓ = mantém notas rápidas


def _run_basic_pitch(audio_path: Path) -> List[Tuple[float, float, int, float]]:
    """Roda basic-pitch e devolve note_events: (start_s, end_s, pitch, amplitude)."""
    os.environ.setdefault("TF_CPP_MIN_LOG_LEVEL", "3")
    from basic_pitch.inference import predict
    _, _, note_events = predict(
        str(audio_path), _onnx_model_path(),
        onset_threshold=ONSET_THRESHOLD,
        frame_threshold=FRAME_THRESHOLD,
        minimum_note_length=MIN_NOTE_LEN_MS,
    )
    # note_events: (start_s, end_s, pitch, amplitude, pitch_bends)
    return [(float(s), float(e), int(p), float(a)) for s, e, p, a, *_ in note_events]


def _salient_by_step(note_events: List[Tuple[float, float, int, float]],
                     bpm: float, n_steps: int
                     ) -> Dict[int, Tuple[int, float, float, float]]:
    """Colapsa para monofônico: 1 nota (a de maior amplitude) por step do grid.

    Retorna {step: (pitch, amp, start_s, end_s)}.
    """
    step_dur = step_duration_seconds(bpm)
    best: Dict[int, Tuple[int, float, float, float]] = {}
    for start_s, end_s, pitch, amp in note_events:
        step = int(round(start_s / step_dur))
        if step < 0 or step >= n_steps:
            continue
        if step not in best or amp > best[step][1]:
            best[step] = (pitch, amp, start_s, end_s)
    return best


def _contour_frets(steps_sorted: List[int],
                   pitch_by_step: Dict[int, int]) -> Dict[int, int]:
    """Mapeia pitch → fret pelo CONTORNO LOCAL.

    Cada nota recebe um fret pela sua posição dentro do registro (min..max) das
    últimas CONTOUR_WINDOW notas. Preserva "subiu de tom → fret maior", mas o
    registro acompanha a melodia — sem o passeio-aleatório que grudava nos
    extremos (Green/Orange) do esquema acumulativo. Resultado: os 5 frets são
    usados conforme a melodia sobe e desce dentro de cada frase.
    """
    from collections import deque
    frets: Dict[int, int] = {}
    win: deque = deque(maxlen=CONTOUR_WINDOW)
    for s in steps_sorted:
        p = pitch_by_step[s]
        win.append(p)
        lo, hi = min(win), max(win)
        if hi == lo:
            frets[s] = START_FRET
        else:
            frets[s] = int(round((p - lo) / (hi - lo) * (N_FRETS - 1)))
    return frets


# ─────────────────────────────────────────────────────────────────────────────
# Polifonia (acordes) + afinamento de densidade
# ─────────────────────────────────────────────────────────────────────────────
def _group_by_step(note_events: List[Tuple[float, float, int, float]],
                   bpm: float, n_steps: int) -> Dict[int, List[Tuple[int, float, float]]]:
    """step do grid → lista de (pitch, amp, dur_sec) das notas que começam nele."""
    from collections import defaultdict
    step_dur = step_duration_seconds(bpm)
    by_step: Dict[int, List[Tuple[int, float, float]]] = defaultdict(list)
    for start_s, end_s, pitch, amp in note_events:
        s = int(round(start_s / step_dur))
        if 0 <= s < n_steps:
            by_step[s].append((pitch, amp, end_s - start_s))
    return by_step


def _metric_strength(step: int) -> int:
    """Força métrica do step no grid de semicolcheias (4/beat, 16/compasso 4-4)."""
    if step % 16 == 0:
        return 4          # downbeat do compasso
    if step % 8 == 0:
        return 3          # meio do compasso
    if step % 4 == 0:
        return 2          # tempo
    if step % 2 == 0:
        return 1          # colcheia
    return 0              # semicolcheia (mais fraca)


def _thin_onsets(by_step: Dict[int, List[Tuple[int, float]]],
                 dur_secs: float, max_onsets_per_sec: float) -> List[int]:
    """Mantém os onsets mais fortes (métrica + amplitude) até o cap de onsets/seg.

    Tempo-independente: o alvo é densidade física (notas/seg), então a redução é
    consistente em qualquer BPM. Músicas abaixo do cap ficam inalteradas.
    """
    steps = list(by_step)
    target = int(round(max_onsets_per_sec * dur_secs))
    if target <= 0 or len(steps) <= target:
        return sorted(steps)
    def score(s: int):
        return (_metric_strength(s), max(t[1] for t in by_step[s]))
    return sorted(sorted(steps, key=score, reverse=True)[:target])


def _chord_frets(notes: List[Tuple[int, float]], lead_fret: int,
                 max_chord: int) -> List[int]:
    """Frets de um onset: acorde COMPACTO de frets adjacentes ancorado no líder.

    O tamanho = nº de tons simultâneos fortes (até max_chord). Frets adjacentes
    (2 notas → líder+vizinho) dão acordes ergonômicos e preservam a distribuição
    equilibrada do líder — sem empurrar os tons pro Orange como o offset por
    intervalo fazia.
    """
    if max_chord <= 1:
        return [lead_fret]
    notes = sorted(notes, key=lambda x: -x[1])     # por amplitude desc
    lead_amp = notes[0][1]
    strong = sum(1 for t in notes[1:max_chord] if t[1] >= CHORD_AMP_FRAC * lead_amp)
    size = 1 + strong
    if size <= 1:
        return [lead_fret]
    start = max(0, min(lead_fret, N_FRETS - size))  # mantém o acorde dentro de 0..4
    return list(range(start, start + size))


def _write_fret_xlsx(out_path: Path, instrument: str,
                     events: List[Tuple[int, List[int], int]],
                     bpm: float) -> int:
    """Escreve a aba resumida de um instrumento de fret, com sustains.

    events: lista de (step, [frets], dur_ticks). Cada fret vira nota MIDI Expert
    (96-100) com a duração dada (tap curto ou sustain).
    """
    from openpyxl import Workbook

    ticks_per_step = TICKS_PER_BEAT // SUBDIV_PER_BEAT
    tempo_us = int(round(60_000_000 / bpm))

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "info"
    ws_info.append(["File Name", "MIDI Type", "Ticks per Beat",
                    "Tempo (µs/beat)", "BPM", "Time Signature"])
    ws_info.append(["notes.mid", "Type 1", TICKS_PER_BEAT,
                    tempo_us, round(bpm, 2), "4/4"])

    ws = wb.create_sheet(instrument)
    ws.append(["#", "Note #", "Note Name", "Channel", "Velocity",
               "Start Tick", "Start (s)", "End Tick", "End (s)",
               "Duration (ticks)", "Duration (s)"])

    rows = []
    for step, frets, dur_ticks in events:
        start_tick = step * ticks_per_step
        end_tick   = start_tick + dur_ticks
        start_s    = (start_tick / TICKS_PER_BEAT) * (tempo_us / 1e6)
        end_s      = (end_tick   / TICKS_PER_BEAT) * (tempo_us / 1e6)
        dur_s      = dur_ticks   / TICKS_PER_BEAT * (tempo_us / 1e6)
        for fr in frets:
            midi = FRET_IDX_TO_MIDI[fr]
            rows.append((midi, midi_to_note_name(midi), 0, 96,
                         start_tick, round(start_s, 4), end_tick, round(end_s, 4),
                         dur_ticks, round(dur_s, 4)))
    rows.sort(key=lambda r: (r[4], r[0]))
    for i, r in enumerate(rows, 1):
        ws.append((i, *r))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return len(rows)


# ─────────────────────────────────────────────────────────────────────────────
# Escrita do xlsx de vocals (pitches reais — não cabe no predictions_to_xlsx)
# ─────────────────────────────────────────────────────────────────────────────
def _write_vocals_xlsx(best: Dict[int, Tuple[int, float, float, float]],
                       steps_sorted: List[int],
                       bpm: float,
                       out_path: Path) -> int:
    from openpyxl import Workbook

    ticks_per_step = TICKS_PER_BEAT // SUBDIV_PER_BEAT
    step_dur = step_duration_seconds(bpm)
    tempo_us = int(round(60_000_000 / bpm))

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "info"
    ws_info.append(["File Name", "MIDI Type", "Ticks per Beat",
                    "Tempo (µs/beat)", "BPM", "Time Signature"])
    ws_info.append(["notes.mid", "Type 1", TICKS_PER_BEAT,
                    tempo_us, round(bpm, 2), "4/4"])

    ws = wb.create_sheet("vocals")
    ws.append(["#", "Note #", "Note Name", "Channel", "Velocity",
               "Start Tick", "Start (s)", "End Tick", "End (s)",
               "Duration (ticks)", "Duration (s)"])

    rows = []
    for s in steps_sorted:
        pitch, _amp, start_s, end_s = best[s]
        pitch = max(VOCAL_MIDI_MIN, min(VOCAL_MIDI_MAX, pitch))
        start_tick = s * ticks_per_step
        dur_steps  = max(1, int(round((end_s - start_s) / step_dur)))
        dur_ticks  = dur_steps * ticks_per_step
        end_tick   = start_tick + dur_ticks
        start_sec  = (start_tick / TICKS_PER_BEAT) * (tempo_us / 1e6)
        end_sec    = (end_tick   / TICKS_PER_BEAT) * (tempo_us / 1e6)
        dur_sec    = dur_ticks   / TICKS_PER_BEAT * (tempo_us / 1e6)
        rows.append((pitch, midi_to_note_name(pitch), 0, 96,
                     start_tick, round(start_sec, 4),
                     end_tick,   round(end_sec, 4),
                     dur_ticks,  round(dur_sec, 4)))
    rows.sort(key=lambda r: r[4])
    for i, r in enumerate(rows, 1):
        ws.append((i, *r))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return len(rows)


# ─────────────────────────────────────────────────────────────────────────────
# API pública — mesma assinatura de modelo_gera_excel.infer
# ─────────────────────────────────────────────────────────────────────────────
def transcribe(audio_path: Path | str,
               bpm: float,
               instrument: str,
               out_xlsx: Path | str,
               n_steps_override: Optional[int] = None,
               **_ignored) -> Path:
    """Transcreve um stem afinado para xlsx parcial via basic-pitch.

    Args:
        audio_path:       stem do instrumento (guitar.ogg / rhythm.ogg / vocals.ogg)
        bpm:              BPM da música
        instrument:       'guitar', 'rhythm' (bass) ou 'vocals'
        out_xlsx:         arquivo .xlsx a gerar (formato resumido)
        n_steps_override: força um nº específico de steps (senão deriva do áudio)
        **_ignored:       aceita model_path/meta_path/device do contrato do CRNN
                          (ignorados — basic-pitch não usa checkpoint)

    Returns:
        Path do .xlsx gerado.
    """
    if instrument == "bass":
        instrument = "rhythm"
    if instrument not in SUPPORTED_INSTRUMENTS:
        raise NotImplementedError(
            f"basic-pitch não suporta {instrument!r}. "
            f"Suportados: {sorted(SUPPORTED_INSTRUMENTS)} (drums → use CRNN)."
        )

    audio_path = Path(audio_path)
    out_xlsx   = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    step_dur   = step_duration_seconds(bpm)
    audio_secs = audio_duration_seconds(audio_path)
    n_steps    = n_steps_override or (int(audio_secs / step_dur) + 1)

    note_events = _run_basic_pitch(audio_path)

    if instrument in ("guitar", "rhythm"):
        # 1) agrupa por step (mantém polifonia)  2) afina densidade pelos onsets
        #    mais fortes  3) líder pelo contorno + acordes  4) sustains nas notas longas
        by_step = _group_by_step(note_events, bpm, n_steps)

        # Seleção de onset: modelo treinado se houver checkpoint; senão heurística.
        import onset_model  # lazy (carrega torch só aqui)
        _model = onset_model.load_model()
        if _model is not None:
            spb = 60.0 / bpm
            tick_events = [(int(p),
                            int(round(st / spb * TICKS_PER_BEAT)),
                            max(1, int(round((en - st) / spb * TICKS_PER_BEAT))),
                            max(1, min(127, int(round(a * 127)))))
                           for st, en, p, a in note_events]
            sel = onset_model.predict_steps(_model, tick_events, instrument, n_steps)
            kept = sorted(set(sel) & set(by_step))    # só steps com pitch (p/ fret/acorde)
            if not kept:                              # fallback se modelo não selecionou nada
                kept = sorted(_thin_onsets(by_step, audio_secs,
                                           MAX_ONSETS_PER_SEC.get(instrument, 3.0)))
        else:
            kept = sorted(_thin_onsets(by_step, audio_secs,
                                       MAX_ONSETS_PER_SEC.get(instrument, 3.0)))
        lead_pitch = {s: max(by_step[s], key=lambda x: x[1])[0] for s in kept}
        lead_fret  = _contour_frets(kept, lead_pitch)
        max_chord  = MAX_CHORD.get(instrument, 1)
        ticks_per_step = TICKS_PER_BEAT // SUBDIV_PER_BEAT

        # duração (em steps) do líder de cada onset; limiar de sustain = percentil
        # por-música (top SUSTAIN_RATE), com piso absoluto → ~taxa humana de sustains
        dur_steps_map = {s: int(round(max(by_step[s], key=lambda x: x[1])[2] / step_dur))
                         for s in kept}
        if dur_steps_map:
            thr = max(SUSTAIN_MIN_STEPS,
                      int(np.quantile(list(dur_steps_map.values()), 1.0 - SUSTAIN_RATE)))
        else:
            thr = SUSTAIN_MIN_STEPS

        events = []
        for idx, s in enumerate(kept):
            frets = _chord_frets(by_step[s], lead_fret[s], max_chord)
            ds = dur_steps_map[s]
            if ds >= thr:
                nxt = kept[idx + 1] if idx + 1 < len(kept) else s + ds + 1
                ds = max(1, min(ds, nxt - s - 1))          # não invade o próximo onset
                dur_ticks = max(TAP_TICKS, ds * ticks_per_step)
            else:
                dur_ticks = TAP_TICKS
            events.append((s, frets, dur_ticks))

        n_events = _write_fret_xlsx(out_xlsx, instrument, events, bpm)
    else:  # vocals — monofônico, pitch real (sem acorde, sem afinar densidade)
        best = _salient_by_step(note_events, bpm, n_steps)
        n_events = _write_vocals_xlsx(best, sorted(best), bpm, out_xlsx)

    log.info("basic-pitch %s: %d nota(s) detectada(s) → %d evento(s) → %s",
             instrument, len(note_events), n_events, out_xlsx.name)
    return out_xlsx


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Transcreve um stem afinado (guitar/bass/vocals) para xlsx via basic-pitch.",
    )
    p.add_argument("--audio",      required=True, type=Path)
    p.add_argument("--bpm",        required=True, type=float)
    p.add_argument("--instrument", required=True, choices=sorted(SUPPORTED_INSTRUMENTS))
    p.add_argument("--out",        required=True, type=Path, dest="out_xlsx")
    p.add_argument("--quiet",      action="store_true")
    return p


def main(argv: Optional[list] = None) -> int:
    args = _build_arg_parser().parse_args(argv)
    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%H:%M:%S",
    )
    out = transcribe(args.audio, args.bpm, args.instrument, args.out_xlsx)
    print(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

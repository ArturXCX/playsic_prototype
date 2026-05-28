"""
organizar_dataset — monta `dados/dataset/` a partir de `dados/pre_dataset/`.

Fluxo:
    1. Processa o conjunto todo usando o método original.
    2. Identifica os MIDIs cujo metadata ficou vazio.
    3. Reprocessa apenas esses MIDIs usando fechamento por note_off.
    4. Se ainda assim o MIDI continuar sem notas, remove a pasta do dataset final
       e remove também da planilha de metadata.

Uso CLI:
    python3 dados/organizar_dataset.py --src dados/pre_dataset/ --dst dados/dataset/

Uso API:
    from dados.organizar_dataset import organize_dataset
    organize_dataset("dados/pre_dataset/", "dados/dataset/")
"""
from __future__ import annotations

import argparse
import configparser
import logging
import os
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import mido
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Constantes
# ─────────────────────────────────────────────────────────────────────────────
OGG_FILES = ["drums.ogg", "guitar.ogg", "rhythm.ogg", "vocals.ogg", "song.ogg"]

MIDI_INFO_HEADERS = [
    "File Name",
    "MIDI Type",
    "Ticks per Beat",
    "Tempo (µs/beat)",
    "BPM",
    "Time Signature",
]

TRACK_MAP = {
    "PART DRUMS":  "drums",
    "PART GUITAR": "guitar",
    "PART BASS":   "rhythm",
    "PART VOCALS": "vocals",
}

NOTE_NAMES = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]

NOTE_HEADERS = [
    "#", "Note #", "Note Name", "Channel", "Velocity",
    "Start Tick", "Start (s)", "End Tick", "End (s)",
    "Duration (ticks)", "Duration (s)",
]

SONG_INI_KEYS = [
    "name", "artist", "album", "charter", "frets", "year", "genre",
    "pro_drums", "song_length", "preview_start_time", "preview_end_time",
    "diff_band", "diff_guitar", "diff_guitarghl", "diff_bass", "diff_bassghl",
    "diff_drums", "diff_drums_real", "diff_keys", "diff_keys_real",
    "diff_vocals", "diff_vocals_harm", "diff_dance", "diff_bass_real",
    "diff_guitar_real", "diff_guitar_coop", "diff_rhythm",
    "diff_drums_real_ps", "diff_keys_real_ps",
    "diff_guitar_pad", "diff_bass_pad", "diff_drums_pad",
    "diff_vocals_pad", "diff_keys_pad",
    "star_power_note", "multiplier_note", "track", "album_track",
    "sysex_slider", "sysex_open_bass",
]

# Mapa de normalização de gêneros para o dataset_metadata.xlsx.
# Gêneros não presentes aqui são mantidos como estão.
GENRE_MAPPING_6: Dict[str, str] = {
    "Pop/Dance/Electronic": "Pop / Dance / Electronic",
    "Dance":                "Pop / Dance / Electronic",
    "Disco/Pop":            "Pop / Dance / Electronic",
    "Disco":                "Pop / Dance / Electronic",
    "Pop":                  "Pop / Dance / Electronic",
    "Art Pop":              "Pop / Dance / Electronic",
    "Synthpop":             "Pop / Dance / Electronic",
    "Synth-Pop":            "Pop / Dance / Electronic",
    "New Wave":             "Pop / Dance / Electronic",
    "Post-Punk Revival":    "Pop / Dance / Electronic",
    "Rock":                 "Rock / Alternative / Indie",
    "J-Rock":               "Rock / Alternative / Indie",
    "Classic Rock":         "Rock / Alternative / Indie",
    "Southern Rock":        "Rock / Alternative / Indie",
    "Pop-Rock":             "Rock / Alternative / Indie",
    "Pop Rock":             "Rock / Alternative / Indie",
    "Pop/Rock":             "Rock / Alternative / Indie",
    "Alternative":          "Rock / Alternative / Indie",
    "Alternative Rock":     "Rock / Alternative / Indie",
    "Indie Rock":           "Rock / Alternative / Indie",
    "Indie":                "Rock / Alternative / Indie",
    "Modern Rock":          "Rock / Alternative / Indie",
    "Grunge":               "Rock / Alternative / Indie",
    "Post-Grunge":          "Rock / Alternative / Indie",
    "Punk":                 "Rock / Alternative / Indie",
    "Punk Rock":            "Rock / Alternative / Indie",
    "Emo":                  "Rock / Alternative / Indie",
    "Psychedelic Rock":     "Rock / Alternative / Indie",
    "Space Rock":           "Rock / Alternative / Indie",
    "Prog":                 "Rock / Alternative / Indie",
    "Progressive":          "Rock / Alternative / Indie",
    "Metal":                "Metal / Hard Rock",
    "Heavy Metal":          "Metal / Hard Rock",
    "Speed Metal":          "Metal / Hard Rock",
    "Power Metal":          "Metal / Hard Rock",
    "Nu-Metal":             "Metal / Hard Rock",
    "Nu Metal":             "Metal / Hard Rock",
    "Alternative Metal":    "Metal / Hard Rock",
    "Groove Metal":         "Metal / Hard Rock",
    "Stoner Rock":          "Metal / Hard Rock",
    "Funk Metal":           "Metal / Hard Rock",
    "Hard Rock":            "Metal / Hard Rock",
    "Arena Rock":           "Metal / Hard Rock",
    "Glam":                 "Metal / Hard Rock",
    "Hip-Hop/Rap":          "Hip-Hop / R&B / Funk",
    "Hiphop/Rap":           "Hip-Hop / R&B / Funk",
    "Hip Hop/Rap":          "Hip-Hop / R&B / Funk",
    "Hip Hop":              "Hip-Hop / R&B / Funk",
    "R&B/Soul/Funk":        "Hip-Hop / R&B / Funk",
    "R&B":                  "Hip-Hop / R&B / Funk",
    "Funk":                 "Hip-Hop / R&B / Funk",
    "Country":              "Country / Roots / World",
    "Reggae/Ska":           "Country / Roots / World",
    "Latin":                "Country / Roots / World",
    "Jazz":                 "Country / Roots / World",
    "Blues":                "Country / Roots / World",
    "Texas Blues":          "Country / Roots / World",
    "Classical":            "Other / Niche",
    "Industrial":           "Other / Niche",
    "Novelty":              "Other / Niche",
    "Other":                "Other / Niche",
}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers MIDI → xlsx resumido
# ─────────────────────────────────────────────────────────────────────────────
def _note_name(n: int) -> str:
    return f"{NOTE_NAMES[n % 12]}{(n // 12) - 1}"


def _ticks_to_seconds(ticks: int, tempo: int, tpb: int) -> float:
    return (ticks / tpb) * (tempo / 1_000_000)


def _get_midi_global_info(mid: mido.MidiFile) -> Tuple[int, int, Tuple[int, int]]:
    tpb = mid.ticks_per_beat
    tempo = 500_000
    time_sig = (4, 4)

    if mid.tracks:
        for msg in mid.tracks[0]:
            if msg.type == "set_tempo":
                tempo = msg.tempo
            elif msg.type == "time_signature":
                time_sig = (msg.numerator, msg.denominator)

    return tpb, tempo, time_sig


def _append_note(
    notes: List[Dict[str, Any]],
    note: int,
    channel: int,
    velocity: int,
    s_tick: int,
    s_sec: float,
    end_tick: int,
    tempo: int,
    tpb: int,
) -> None:
    dur_ticks = end_tick - s_tick
    dur_sec = _ticks_to_seconds(dur_ticks, tempo, tpb)

    notes.append({
        "note":           note,
        "note_name":      _note_name(note),
        "channel":        channel,
        "velocity":       velocity,
        "start_tick":     s_tick,
        "start_sec":      round(s_sec, 4),
        "end_tick":       end_tick,
        "end_sec":        round(s_sec + dur_sec, 4),
        "duration_ticks": dur_ticks,
        "duration_sec":   round(dur_sec, 4),
    })


def _parse_midi_original(path: Path) -> Dict[str, Any]:
    """
    Parser original.

    Fecha nota apenas quando encontra:
        note_on velocity == 0

    Este método é mantido para a primeira passada do dataset.
    """
    mid = mido.MidiFile(str(path))
    tpb, tempo, time_sig = _get_midi_global_info(mid)

    tracks: List[Dict[str, Any]] = []

    for track in mid.tracks[1:]:
        notes: List[Dict[str, Any]] = []
        abs_tick = 0
        open_notes: Dict[int, Tuple[int, float, int, int]] = {}

        for msg in track:
            abs_tick += msg.time

            if msg.type == "note_on" and msg.velocity > 0:
                open_notes[msg.note] = (
                    abs_tick,
                    _ticks_to_seconds(abs_tick, tempo, tpb),
                    msg.velocity,
                    msg.channel,
                )

            elif msg.type == "note_on" and msg.velocity == 0:
                if msg.note in open_notes:
                    s_tick, s_sec, vel, ch = open_notes.pop(msg.note)

                    _append_note(
                        notes=notes,
                        note=msg.note,
                        channel=ch,
                        velocity=vel,
                        s_tick=s_tick,
                        s_sec=s_sec,
                        end_tick=abs_tick,
                        tempo=tempo,
                        tpb=tpb,
                    )

        tracks.append({"name": track.name, "notes": notes})

    return {
        "filename":       os.path.basename(str(path)),
        "type":           mid.type,
        "ticks_per_beat": tpb,
        "tempo_us":       tempo,
        "bpm":            round(60_000_000 / tempo, 2),
        "time_signature": f"{time_sig[0]}/{time_sig[1]}",
        "tracks":         tracks,
    }


def _parse_midi_com_note_off(path: Path) -> Dict[str, Any]:
    """
    Parser alternativo.

    Fecha nota quando encontra:
        - note_off
        - note_on velocity == 0

    Usado apenas na segunda passada, para recuperar MIDIs cujo metadata
    ficou vazio no parser original.
    """
    mid = mido.MidiFile(str(path))
    tpb, tempo, time_sig = _get_midi_global_info(mid)

    tracks: List[Dict[str, Any]] = []

    for track in mid.tracks[1:]:
        notes: List[Dict[str, Any]] = []
        abs_tick = 0

        open_notes: Dict[Tuple[int, int], List[Tuple[int, float, int]]] = {}

        for msg in track:
            abs_tick += msg.time

            if msg.type == "note_on" and msg.velocity > 0:
                key = (msg.channel, msg.note)

                if key not in open_notes:
                    open_notes[key] = []

                open_notes[key].append((
                    abs_tick,
                    _ticks_to_seconds(abs_tick, tempo, tpb),
                    msg.velocity,
                ))

            elif msg.type == "note_off" or (msg.type == "note_on" and msg.velocity == 0):
                key = (msg.channel, msg.note)
                stack = open_notes.get(key)

                if stack:
                    s_tick, s_sec, vel = stack.pop()

                    if not stack:
                        del open_notes[key]

                    _append_note(
                        notes=notes,
                        note=msg.note,
                        channel=msg.channel,
                        velocity=vel,
                        s_tick=s_tick,
                        s_sec=s_sec,
                        end_tick=abs_tick,
                        tempo=tempo,
                        tpb=tpb,
                    )

        tracks.append({"name": track.name, "notes": notes})

    return {
        "filename":       os.path.basename(str(path)),
        "type":           mid.type,
        "ticks_per_beat": tpb,
        "tempo_us":       tempo,
        "bpm":            round(60_000_000 / tempo, 2),
        "time_signature": f"{time_sig[0]}/{time_sig[1]}",
        "tracks":         tracks,
    }


def _all_tracks_have_zero_notes(data: Dict[str, Any]) -> bool:
    return not any(len(t.get("notes", [])) > 0 for t in data.get("tracks", []))


def midi_to_xlsx_resumido(
    midi_path: Path,
    out_xlsx: Path,
    metodo: str = "original",
) -> Tuple[Dict[str, Any], bool]:
    """
    Converte notes.mid → notes.xlsx resumido.

    metodo:
        "original"  -> fecha notas só com note_on velocity == 0
        "note_off"  -> fecha notas com note_off ou note_on velocity == 0

    Retorna:
        (midi_info, all_tracks_zero_notes)
    """
    if metodo == "original":
        data = _parse_midi_original(midi_path)
    elif metodo == "note_off":
        data = _parse_midi_com_note_off(midi_path)
    else:
        raise ValueError(f"Método inválido: {metodo}")

    all_zero = _all_tracks_have_zero_notes(data)

    wb = openpyxl.Workbook()
    ws_info = wb.active
    ws_info.title = "info"

    ws_info.append(MIDI_INFO_HEADERS)

    info_values = [
        data["filename"],
        f"Type {data['type']}",
        data["ticks_per_beat"],
        data["tempo_us"],
        data["bpm"],
        data["time_signature"],
    ]

    ws_info.append(info_values)
    midi_info = dict(zip(MIDI_INFO_HEADERS, info_values))

    for track in data["tracks"]:
        raw_name = track["name"].strip()
        base_name = raw_name.replace(" (0)", "").strip()
        sheet_name = TRACK_MAP.get(base_name)

        if sheet_name is None:
            continue

        ws = wb.create_sheet(title=sheet_name)
        ws.append(NOTE_HEADERS)

        for i, n in enumerate(track["notes"], 1):
            ws.append([
                i,
                n["note"],
                n["note_name"],
                n["channel"],
                n["velocity"],
                n["start_tick"],
                n["start_sec"],
                n["end_tick"],
                n["end_sec"],
                n["duration_ticks"],
                n["duration_sec"],
            ])

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)

    return midi_info, all_zero


# ─────────────────────────────────────────────────────────────────────────────
# Leitura do song.ini
# ─────────────────────────────────────────────────────────────────────────────
def parse_song_ini(song_ini_path: Path) -> Dict[str, Optional[str]]:
    values: Dict[str, Optional[str]] = {k: None for k in SONG_INI_KEYS}

    if not song_ini_path.exists():
        return values

    text = None

    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = song_ini_path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue

    if text is None:
        return values

    parser = configparser.ConfigParser(interpolation=None)
    parser.optionxform = str

    try:
        parser.read_string(text)
    except configparser.Error:
        for line in text.splitlines():
            clean = line.strip()

            if not clean or clean.startswith(("#", ";", "[")) or "=" not in clean:
                continue

            key, value = clean.split("=", 1)
            key = key.strip()

            if key in values:
                values[key] = value.strip()

        return values

    if not parser.has_section("song"):
        return values

    section = parser["song"]
    lower_to_original = {k.lower(): k for k in section.keys()}

    for key in SONG_INI_KEYS:
        if key in section:
            values[key] = section.get(key)
        elif key.lower() in lower_to_original:
            values[key] = section.get(lower_to_original[key.lower()])

    return values


# ─────────────────────────────────────────────────────────────────────────────
# dataset_metadata.xlsx
# ─────────────────────────────────────────────────────────────────────────────
def _save_metadata_xlsx(rows: List[Dict[str, Any]], out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "metadata"

    headers = ["Folder Name"] + OGG_FILES + MIDI_INFO_HEADERS + SONG_INI_KEYS
    ws.append(headers)

    genre_idx = headers.index("genre") if "genre" in headers else -1

    for row in rows:
        row_values = [row.get(h) for h in headers]
        if genre_idx >= 0 and row_values[genre_idx] is not None:
            row_values[genre_idx] = GENRE_MAPPING_6.get(
                str(row_values[genre_idx]), row_values[genre_idx]
            )
        ws.append(row_values)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0

        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 2, 10),
            35,
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers de organização
# ─────────────────────────────────────────────────────────────────────────────
def _copy_assets_and_build_base_metadata(src_folder: Path, dst_folder: Path) -> Dict[str, Any]:
    folder_name = src_folder.name
    metadata_row: Dict[str, Any] = {"Folder Name": folder_name}

    for ogg in OGG_FILES:
        src_ogg = src_folder / ogg
        dst_ogg = dst_folder / ogg

        if src_ogg.exists():
            shutil.copy2(src_ogg, dst_ogg)
            log.info("  ✔ %s", ogg)

        elif ogg == "rhythm.ogg":
            src_bass = src_folder / "bass.ogg"

            if src_bass.exists():
                shutil.copy2(src_bass, dst_ogg)
                log.info("  ✔ bass.ogg → rhythm.ogg")

        metadata_row[ogg] = int(dst_ogg.exists())

    src_album = src_folder / "album.png"

    if src_album.exists():
        shutil.copy2(src_album, dst_folder / "album.png")

    return metadata_row


def _complete_metadata_row(
    src_folder: Path,
    dst_folder: Path,
    midi_info: Dict[str, Any],
) -> Dict[str, Any]:
    folder_name = src_folder.name
    row: Dict[str, Any] = {"Folder Name": folder_name}

    for ogg in OGG_FILES:
        row[ogg] = int((dst_folder / ogg).exists())

    row.update(midi_info)
    row.update(parse_song_ini(src_folder / "song.ini"))

    return row


# ─────────────────────────────────────────────────────────────────────────────
# API pública
# ─────────────────────────────────────────────────────────────────────────────
def organize_dataset(
    src: Path | str,
    dst: Path | str,
    metadata_xlsx: Optional[Path | str] = None,
) -> Dict[str, int]:
    """
    Pipeline completo.

    Primeira passada:
        - Copia arquivos.
        - Converte notes.mid com o parser original.
        - Se o MIDI ficar sem notas, não grava metadata ainda.

    Segunda passada:
        - Reprocessa somente os MIDIs vazios usando fechamento por note_off.
        - Se recuperar notas, atualiza notes.xlsx e metadata.
        - Se continuar vazio, remove a pasta do dataset final.
    """
    src = Path(src)
    dst = Path(dst)
    dst.mkdir(parents=True, exist_ok=True)

    metadata_xlsx = Path(metadata_xlsx) if metadata_xlsx else dst / "dataset_metadata.xlsx"

    folders = sorted([f for f in src.iterdir() if f.is_dir()])
    log.info("Pastas a processar: %d", len(folders))

    ok = 0
    err = 0
    empty_first_pass = 0
    recovered = 0
    removed = 0

    rows_by_folder: Dict[str, Dict[str, Any]] = {}
    empty_candidates: List[Path] = []

    # ─────────────────────────────────────────────────────────────────────
    # Primeira passada: método original
    # ─────────────────────────────────────────────────────────────────────
    log.info("Primeira passada: processamento original")

    for src_folder in folders:
        folder_name = src_folder.name
        dst_folder = dst / folder_name
        dst_folder.mkdir(parents=True, exist_ok=True)

        log.info("[%s]", folder_name)

        metadata_row = _copy_assets_and_build_base_metadata(src_folder, dst_folder)

        mid_path = src_folder / "notes.mid"
        midi_info: Dict[str, Any] = {h: None for h in MIDI_INFO_HEADERS}
        all_zero = False

        if mid_path.exists():
            try:
                midi_info, all_zero = midi_to_xlsx_resumido(
                    midi_path=mid_path,
                    out_xlsx=dst_folder / "notes.xlsx",
                    metodo="original",
                )
                ok += 1

            except Exception as exc:
                log.warning("  ✘ erro processando notes.mid no método original: %s", exc)
                err += 1

        else:
            log.warning("  – notes.mid ausente, pulando MIDI")

        if all_zero:
            empty_first_pass += 1
            empty_candidates.append(src_folder)
            log.info("  – metadata vazio na primeira passada")
            continue

        metadata_row.update(midi_info)
        metadata_row.update(parse_song_ini(src_folder / "song.ini"))
        rows_by_folder[folder_name] = metadata_row

    # ─────────────────────────────────────────────────────────────────────
    # Segunda passada: recuperar vazios com note_off
    # ─────────────────────────────────────────────────────────────────────
    if empty_candidates:
        log.info(
            "Segunda passada: reprocessando %d MIDI(s) vazios com note_off",
            len(empty_candidates),
        )

    for src_folder in empty_candidates:
        folder_name = src_folder.name
        dst_folder = dst / folder_name
        mid_path = src_folder / "notes.mid"

        log.info("[%s] segunda passada com note_off", folder_name)

        if not mid_path.exists():
            log.warning("  – notes.mid ausente; removendo do dataset final")
            if dst_folder.exists():
                shutil.rmtree(dst_folder)
            removed += 1
            continue

        try:
            midi_info, still_zero = midi_to_xlsx_resumido(
                midi_path=mid_path,
                out_xlsx=dst_folder / "notes.xlsx",
                metodo="note_off",
            )

        except Exception as exc:
            log.warning("  ✘ erro no reprocessamento com note_off: %s", exc)
            err += 1

            if dst_folder.exists():
                shutil.rmtree(dst_folder)

            removed += 1
            continue

        if still_zero:
            log.info("  – ainda vazio após note_off; removendo do dataset final")

            if dst_folder.exists():
                shutil.rmtree(dst_folder)

            removed += 1
            continue

        row = _complete_metadata_row(
            src_folder=src_folder,
            dst_folder=dst_folder,
            midi_info=midi_info,
        )

        rows_by_folder[folder_name] = row
        recovered += 1

        log.info("  ✔ recuperado com note_off")

    # Mantém a ordem original das pastas, excluindo as removidas
    rows = [
        rows_by_folder[f.name]
        for f in folders
        if f.name in rows_by_folder
    ]

    _save_metadata_xlsx(rows, metadata_xlsx)

    log.info("dataset_metadata.xlsx salvo em %s (%d linhas)", metadata_xlsx, len(rows))

    return {
        "ok": ok,
        "err": err,
        "empty": removed,
        "empty_first_pass": empty_first_pass,
        "recovered": recovered,
        "removed": removed,
        "metadata_rows": len(rows),
    }


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────
def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Monta dados/dataset/ a partir de dados/pre_dataset/ + planilha consolidada.",
    )

    p.add_argument(
        "--src",
        required=True,
        type=Path,
        help="Pasta de entrada, ex: dados/pre_dataset/",
    )

    p.add_argument(
        "--dst",
        required=True,
        type=Path,
        help="Pasta de destino, ex: dados/dataset/",
    )

    p.add_argument(
        "--metadata-xlsx",
        type=Path,
        default=None,
        help="Caminho do dataset_metadata.xlsx. Default: <dst>/dataset_metadata.xlsx",
    )

    p.add_argument("--quiet", action="store_true")

    return p


def main(argv: Optional[list] = None) -> int:
    args = _build_arg_parser().parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%H:%M:%S",
    )

    if not args.src.is_dir():
        raise FileNotFoundError(f"Pasta src não existe: {args.src}")

    stats = organize_dataset(
        src=args.src,
        dst=args.dst,
        metadata_xlsx=args.metadata_xlsx,
    )

    print("─" * 70)
    print(f"✓ {stats['ok']} MIDI(s) processado(s) na primeira passada")
    print(f"↻ {stats['empty_first_pass']} metadata vazio(s) detectado(s)")
    print(f"✓ {stats['recovered']} recuperado(s) com note_off")
    print(f"🗑 {stats['removed']} removido(s) do dataset final")
    print(f"✘ {stats['err']} erro(s)")
    print(f"📄 {stats['metadata_rows']} linha(s) finais no dataset_metadata.xlsx")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
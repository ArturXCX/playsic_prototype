"""
notes_xlsx — leitura/escrita do notes.xlsx no formato resumido.

Formato resumido (mesmo gerado por `dados/organizar_dataset.py`):
    Aba 'info'    — 1 cabeçalho + 1 linha:
        File Name | MIDI Type | Ticks per Beat | Tempo (µs/beat) | BPM | Time Signature
    Abas 'drums'/'guitar'/'rhythm'/'vocals' — cabeçalho com 11 colunas + N linhas:
        # | Note # | Note Name | Channel | Velocity |
        Start Tick | Start (s) | End Tick | End (s) |
        Duration (ticks) | Duration (s)

Funções:
    - `parse_info_sheet`      → metadados globais
    - `parse_drum_events`     → eventos da aba 'drums' filtrados nas lanes RB
    - `predictions_to_xlsx`   → grava predições binárias num xlsx parcial
    - `events_to_target_matrix` → constrói o target binário [n_steps, n_lanes]
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Tuple

import numpy as np
from openpyxl import Workbook, load_workbook


NOTE_NAMES = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]


def midi_to_note_name(n: int) -> str:
    return f"{NOTE_NAMES[n % 12]}{(n // 12) - 1}"


# ─────────────────────────────────────────────────────────────────────────────
# Leitura
# ─────────────────────────────────────────────────────────────────────────────
def parse_info_sheet(xlsx_path: Path | str) -> Dict[str, Any]:
    """Lê a aba 'info' e retorna {filename, bpm, ticks_per_beat, tempo_us, time_signature}."""
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        ws = wb["info"]
        rows = list(ws.iter_rows(values_only=True, max_row=2))
        info = dict(zip(rows[0], rows[1]))
    finally:
        wb.close()
    return {
        "filename":       info.get("File Name"),
        "bpm":            float(info["BPM"]),
        "ticks_per_beat": int(info["Ticks per Beat"]),
        "tempo_us":       int(info.get("Tempo (µs/beat)", 500_000)),
        "time_signature": str(info.get("Time Signature", "4/4")),
    }


def parse_drum_events(xlsx_path: Path | str,
                      lanes_map: Dict[int, int],
                      gameplay_max_midi: int = 36) -> Dict[str, Any]:
    """Lê a aba 'drums' e devolve eventos válidos + duração.

    Filtra:
        - MIDI ≥ `gameplay_max_midi` (mirror visual)
        - MIDI fora de `lanes_map`

    Returns:
        {
          'events': List[(start_tick, lane_idx)],
          'duration_ticks': último end_tick observado,
        }
    """
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    events: List[Tuple[int, int]] = []
    max_end_tick = 0
    try:
        ws = wb["drums"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                continue
            midi_num = int(row[1])
            if midi_num >= gameplay_max_midi or midi_num not in lanes_map:
                continue
            start_tick = int(row[5])
            end_tick   = int(row[7]) if row[7] is not None else start_tick
            events.append((start_tick, lanes_map[midi_num]))
            max_end_tick = max(max_end_tick, end_tick)
    finally:
        wb.close()
    return {"events": events, "duration_ticks": max_end_tick}


def _parse_lane_events(xlsx_path: Path | str,
                       sheet_name: str,
                       lanes_map: Dict[int, int]) -> Dict[str, Any]:
    """Helper genérico: lê uma aba e mantém só MIDI notes presentes em lanes_map."""
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    events: List[Tuple[int, int]] = []
    max_end_tick = 0
    try:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                continue
            midi_num = int(row[1])
            if midi_num not in lanes_map:
                continue
            start_tick = int(row[5])
            end_tick   = int(row[7]) if row[7] is not None else start_tick
            events.append((start_tick, lanes_map[midi_num]))
            max_end_tick = max(max_end_tick, end_tick)
    except KeyError:
        pass   # aba ausente (música sem aquele instrumento)
    finally:
        wb.close()
    return {"events": events, "duration_ticks": max_end_tick}


def parse_rhythm_events(xlsx_path: Path | str,
                        lanes_map: Dict[int, int]) -> Dict[str, Any]:
    """Lê a aba 'rhythm' (PART BASS) e devolve eventos válidos + duração.

    Para bass/rhythm guitar, as notas Expert do Clone Hero são 96-100.
    Notas fora de `lanes_map` (HOPO markers, star power, dificuldades
    inferiores, etc.) são ignoradas automaticamente.
    """
    return _parse_lane_events(xlsx_path, "rhythm", lanes_map)


def parse_guitar_events(xlsx_path: Path | str,
                        lanes_map: Dict[int, int]) -> Dict[str, Any]:
    """Lê a aba 'guitar' (PART GUITAR) e devolve eventos válidos + duração.

    Para guitar lead, as notas Expert do Clone Hero são 96-100. Mesmo
    schema que parse_rhythm_events, apenas trocando o nome da aba.
    """
    return _parse_lane_events(xlsx_path, "guitar", lanes_map)


def parse_vocals_events(xlsx_path: Path | str,
                        lanes_map: Dict[int, int],
                        midi_min: int = 36,
                        midi_max: int = 84) -> Dict[str, Any]:
    """Lê a aba 'vocals' (PART VOCALS) e devolve eventos de vocal activity.

    v1 (single-lane onset detector):
        - Qualquer nota com pitch em [midi_min, midi_max] vira evento na lane 0.
        - Pitches fora do range (markers de phrase, talk-overs muito graves)
          são ignorados.
        - O parâmetro `lanes_map` é aceito por compatibilidade com a interface
          de preprocess_song mas é IGNORADO (v1 só tem 1 lane).
    """
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    events: List[Tuple[int, int]] = []
    max_end_tick = 0
    try:
        ws = wb["vocals"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] is None:
                continue
            midi_num = int(row[1])
            if midi_num < midi_min or midi_num > midi_max:
                continue
            start_tick = int(row[5])
            end_tick   = int(row[7]) if row[7] is not None else start_tick
            events.append((start_tick, 0))   # tudo cai na única lane
            max_end_tick = max(max_end_tick, end_tick)
    except KeyError:
        pass   # música sem vocals chart
    finally:
        wb.close()
    return {"events": events, "duration_ticks": max_end_tick}


def events_to_target_matrix(events: List[Tuple[int, int]],
                            n_steps: int,
                            ticks_per_step: int,
                            n_lanes: int) -> np.ndarray:
    """Snap-to-grid de eventos → matriz binária [n_steps, n_lanes]."""
    target = np.zeros((n_steps, n_lanes), dtype=np.float32)
    for tick, lane in events:
        step = int(round(tick / ticks_per_step))
        if 0 <= step < n_steps:
            target[step, lane] = 1.0
    return target


# ─────────────────────────────────────────────────────────────────────────────
# Escrita: predições do modelo → xlsx parcial
# ─────────────────────────────────────────────────────────────────────────────
def predictions_to_xlsx(preds: np.ndarray,
                        instrument: str,
                        idx_to_midi: Dict[int, int],
                        bpm: float,
                        ticks_per_beat: int,
                        out_path: Path | str,
                        subdiv_per_beat: int = 4,
                        velocity: int = 96,
                        channel: int = 0,
                        duration_ticks: int = 60) -> int:
    """Salva um xlsx parcial com abas 'info' + `<instrument>` (formato resumido).

    Args:
        preds:          matriz binária [n_steps, n_lanes]
        instrument:     'drums' / 'guitar' / 'rhythm' / 'vocals'
        idx_to_midi:    map lane_idx → MIDI note (inverso de LANES)
        bpm:            BPM da música (usado no cálculo do tempo µs)
        ticks_per_beat: padrão 480
        out_path:       arquivo .xlsx a escrever
        velocity, channel, duration_ticks: defaults para todos os eventos

    Returns:
        Número de eventos escritos.
    """
    out_path = Path(out_path)
    ticks_per_step = ticks_per_beat // subdiv_per_beat
    tempo_us = int(round(60_000_000 / bpm))

    wb = Workbook()

    # ── info ──────────────────────────────────────────
    ws_info = wb.active
    ws_info.title = "info"
    ws_info.append(["File Name", "MIDI Type", "Ticks per Beat",
                    "Tempo (µs/beat)", "BPM", "Time Signature"])
    ws_info.append(["notes.mid", "Type 1", ticks_per_beat,
                    tempo_us, round(bpm, 2), "4/4"])

    # ── instrumento ───────────────────────────────────
    ws = wb.create_sheet(instrument)
    ws.append(["#", "Note #", "Note Name", "Channel", "Velocity",
               "Start Tick", "Start (s)", "End Tick", "End (s)",
               "Duration (ticks)", "Duration (s)"])

    rows = []
    n_lanes = preds.shape[1]
    for step in range(preds.shape[0]):
        for lane in range(n_lanes):
            if preds[step, lane] <= 0.5:
                continue
            midi_num   = idx_to_midi[lane]
            start_tick = step * ticks_per_step
            end_tick   = start_tick + duration_ticks
            start_s    = (start_tick / ticks_per_beat) * (tempo_us / 1e6)
            end_s      = (end_tick   / ticks_per_beat) * (tempo_us / 1e6)
            dur_s      = duration_ticks / ticks_per_beat * (tempo_us / 1e6)
            rows.append((midi_num, midi_to_note_name(midi_num),
                         channel, velocity,
                         start_tick, round(start_s, 4),
                         end_tick,   round(end_s, 4),
                         duration_ticks, round(dur_s, 4)))
    rows.sort(key=lambda r: (r[4], r[0]))
    for i, r in enumerate(rows, 1):
        ws.append((i, *r))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return len(rows)

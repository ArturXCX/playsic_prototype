"""
comparar_transcricao — diagnóstico: basic-pitch vs chart real (ground truth).

Para UMA música do dataset (que tem notes.xlsx = chart real), compara:
  1. Ground truth Expert  — o chart humano (notas, densidade, diversidade de frets)
  2. basic-pitch          — cru → após colapso monofônico → após contorno→fret
  3. Dificuldades         — autoria humana (Easy/Med/Hard/Expert no xlsx) vs a
                            redução algorítmica do excel_to_midi

Objetivo: localizar de onde vem a simplicidade ("poucas notas, pouca diversidade").

Uso:
    python treinamento/comparar_transcricao.py --song "3 Doors Down - Kryptonite"
"""
from __future__ import annotations

import argparse
import collections
import os
import sys
from pathlib import Path

os.environ.setdefault("TF_CPP_MIN_LOG_LEVEL", "3")

try:                                   # console Windows é cp1252; força UTF-8
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "treinamento"))
sys.path.insert(0, str(ROOT / "processamento" / "audio"))
sys.path.insert(0, str(ROOT / "processamento" / "midi_excel"))

from openpyxl import load_workbook
from audio_features import SUBDIV_PER_BEAT, audio_duration_seconds, step_duration_seconds
from notes_xlsx import (parse_info_sheet, parse_guitar_events, parse_rhythm_events,
                        parse_vocals_events, parse_drum_events)
from guitar_crnn import FRETS as GUI_FRETS, FRET_NAMES as GUI_NAMES
from bass_crnn import FRETS as BASS_FRETS, FRET_NAMES as BASS_NAMES
from drum_crnn import LANES as DRUM_LANES, GAMEPLAY_MAX_MIDI
import transcreve_basic_pitch as bp
import excel_to_midi as e2m


def _bar(frac, width=24):
    return "█" * int(round(frac * width))


def _dist_line(counter, names):
    total = sum(counter.values()) or 1
    parts = []
    for i, nm in enumerate(names):
        pct = 100.0 * counter.get(i, 0) / total
        parts.append(f"{nm[:3]}={pct:4.1f}%")
    return "  ".join(parts)


def gt_fret_stats(events):
    """events=[(tick, fret_idx)] -> (n_notas, n_onsets, chord_rate, Counter(fret))."""
    by_tick = collections.defaultdict(list)
    for tick, idx in events:
        by_tick[tick].append(idx)
    n_notes  = len(events)
    n_onsets = len(by_tick)
    n_chords = sum(1 for v in by_tick.values() if len(v) > 1)
    fret_ct  = collections.Counter(idx for _t, idx in events)
    chord_rate = n_chords / n_onsets if n_onsets else 0.0
    return n_notes, n_onsets, chord_rate, fret_ct


def basic_pitch_stats(stem_path, bpm, instrument="guitar"):
    """Espelha o pipeline real (afinamento + acordes). Devolve dict de stats."""
    dur = audio_duration_seconds(stem_path)
    step_dur = step_duration_seconds(bpm)
    n_steps = int(dur / step_dur) + 1
    raw = bp._run_basic_pitch(Path(stem_path))

    if instrument == "vocals":
        best = bp._salient_by_step(raw, bpm, n_steps)
        pitch_ct = collections.Counter(p for p, *_ in best.values())
        return dict(raw=len(raw), onsets=len(best), notes=len(best),
                    chord_rate=0.0, fret_ct=collections.Counter(),
                    pitch_ct=pitch_ct, dur=dur)

    by_step = bp._group_by_step(raw, bpm, n_steps)
    kept    = bp._thin_onsets(by_step, dur, bp.MAX_ONSETS_PER_SEC.get(instrument, 3.0))
    lead_pitch = {s: max(by_step[s], key=lambda x: x[1])[0] for s in kept}
    lead_fret  = bp._contour_frets(kept, lead_pitch)
    max_chord = bp.MAX_CHORD.get(instrument, 1)
    fret_ct = collections.Counter()
    n_notes = n_chords = 0
    for s in kept:
        frets = bp._chord_frets(by_step[s], lead_fret[s], max_chord)
        for f in frets:
            fret_ct[f] += 1
        n_notes += len(frets)
        if len(frets) > 1:
            n_chords += 1
    chord_rate = n_chords / len(kept) if kept else 0.0
    return dict(raw=len(raw), onsets=len(kept), notes=n_notes,
                chord_rate=chord_rate, fret_ct=fret_ct,
                pitch_ct=collections.Counter(), dur=dur)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--song", default="3 Doors Down - Kryptonite")
    args = ap.parse_args()

    song_dir = ROOT / "dados" / "dataset" / args.song
    xlsx = song_dir / "notes.xlsx"
    if not xlsx.exists():
        raise SystemExit(f"notes.xlsx não encontrado: {xlsx}")

    info = parse_info_sheet(xlsx)
    bpm = info["bpm"]
    print(f"\n{'='*78}\n  {args.song}    BPM={bpm}\n{'='*78}")

    # ── Comparação por instrumento afinado ────────────────────────────────────
    afinados = [
        ("guitar", "guitar.ogg", parse_guitar_events, GUI_FRETS, GUI_NAMES),
        ("rhythm", "rhythm.ogg", parse_rhythm_events, BASS_FRETS, BASS_NAMES),
    ]
    for inst, stem, parse_fn, fretmap, names in afinados:
        ev = parse_fn(xlsx, fretmap)["events"]
        gt_notes, gt_onsets, gt_chord, gt_fret = gt_fret_stats(ev)
        st  = basic_pitch_stats(song_dir / stem, bpm, inst)
        dur = st["dur"]
        gt_nps = gt_onsets / dur if dur else 0
        bp_nps = st["onsets"] / dur if dur else 0
        print(f"\n── {inst.upper()} ──  (duração {dur:.0f}s)")
        print(f"  GROUND TRUTH (Expert):  {gt_onsets:4d} onsets  ({gt_notes} notas, "
              f"chord_rate={gt_chord:.0%})   {gt_nps:.2f} onsets/s")
        print(f"     frets: {_dist_line(gt_fret, names)}")
        print(f"  BASIC-PITCH:            {st['onsets']:4d} onsets  "
              f"({st['notes']} notas, chord_rate={st['chord_rate']:.0%}; cru={st['raw']})   "
              f"{bp_nps:.2f} onsets/s")
        print(f"     frets: {_dist_line(st['fret_ct'], names)}")
        print(f"  → densidade ours/GT = {bp_nps/gt_nps:.0%}   "
              f"frets usados: ours={len([k for k,v in st['fret_ct'].items() if v])}/5  "
              f"GT={len([k for k,v in gt_fret.items() if v])}/5")

    # ── Vocals (pitch real) ───────────────────────────────────────────────────
    ev = parse_vocals_events(xlsx, {-1: 0})["events"]
    gt_voc = len(ev)
    stv = basic_pitch_stats(song_dir / "vocals.ogg", bpm, "vocals")
    dur_v = stv["dur"]
    print(f"\n── VOCALS ──  (duração {dur_v:.0f}s)")
    print(f"  GROUND TRUTH:  {gt_voc:4d} notas   {gt_voc/dur_v:.2f} notas/s")
    print(f"  BASIC-PITCH:   {stv['notes']:4d} notas (cru={stv['raw']})   "
          f"{stv['notes']/dur_v:.2f} notas/s   pitches únicos={len(stv['pitch_ct'])}")

    # ── Drums (referência; usa CRNN, não basic-pitch) ─────────────────────────
    ev_d = parse_drum_events(xlsx, DRUM_LANES, GAMEPLAY_MAX_MIDI)["events"]
    print(f"\n── DRUMS (ref; via CRNN) ──  GROUND TRUTH: {len(ev_d)} eventos gameplay")

    # ── Dificuldades: autoria humana vs nossa redução (guitar) ────────────────
    print(f"\n{'='*78}\n  DIFICULDADES — guitar: autoria humana (no xlsx) vs nossa redução\n{'='*78}")
    ws = load_workbook(xlsx, read_only=True)["guitar"]
    human = collections.Counter()
    ranges = {"Easy": (60, 64), "Medium": (72, 76), "Hard": (84, 88), "Expert": (96, 100)}
    for r in ws.iter_rows(min_row=2, values_only=True):
        n = r[1]
        if n is None:
            continue
        for lvl, (lo, hi) in ranges.items():
            if lo <= n <= hi:
                human[lvl] += 1
    # nossa redução a partir do Expert real
    expert_notes = []
    ticks_per_step = info["ticks_per_beat"] // SUBDIV_PER_BEAT
    for tick, idx in parse_guitar_events(xlsx, GUI_FRETS)["events"]:
        midi = 96 + idx
        expert_notes.append(dict(note=midi, channel=0, velocity=100,
                                 start_tick=tick, end_tick=tick + 60))
    reduced = e2m._reduce_to_difficulties(expert_notes, info["ticks_per_beat"],
                                          e2m._GUITAR_FRET_TO_DIFFICULTIES,
                                          e2m._fret_chord_priority)
    ours = collections.Counter()
    for n in reduced:
        for lvl, (lo, hi) in ranges.items():
            if lo <= n["note"] <= hi:
                ours[lvl] += 1
    print(f"  {'nível':<8} {'humano':>8} {'nossa redução':>15}")
    print("  " + "-" * 33)
    for lvl in ("Easy", "Medium", "Hard", "Expert"):
        print(f"  {lvl:<8} {human[lvl]:>8} {ours[lvl]:>15}")
    print(f"\n  (Expert humano inclui acordes/sustains; nossa redução parte do Expert REAL parseado)")


if __name__ == "__main__":
    main()
